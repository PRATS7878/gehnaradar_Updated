"""Build the daily Excel report from data/catalog.json.

Sheets: Summary | All Items | Rising Searches | Category Interest | Source Status
Output: output/GehnaRadar_YYYY-MM-DD.xlsx (also copied to latest.xlsx)
"""
import json
import shutil
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA, OUT = ROOT / "data", ROOT / "output"
OUT.mkdir(exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="2B2140")
HEADER_FONT = Font(color="F5C86A", bold=True, size=11)
LINK_FONT = Font(color="1155CC", underline="single")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def run():
    cat = json.loads((DATA / "catalog.json").read_text())
    status = json.loads((DATA / "source_status.json").read_text()) \
        if (DATA / "source_status.json").exists() else {}
    items = cat["items"]
    labels, markets = cat["categories"], cat["markets"]
    today = date.today().isoformat()

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.append(["GehnaRadar daily trend report", today])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total items in catalog", len(items)])
    ws.append(["New today", sum(1 for i in items if i["first_seen"] == today)])
    ws.append([])
    ws.append(["Items by category", ""])
    for slug, n in Counter(i["category"] for i in items).most_common():
        ws.append([labels.get(slug, slug), n])
    ws.append([])
    ws.append(["Items by source", ""])
    for src, n in Counter(i["source"] for i in items).most_common():
        ws.append([src, n])
    autosize(ws, [34, 14])

    # ---- All Items ----
    ws = wb.create_sheet("All Items")
    cols = ["First seen", "Category", "Title", "Source", "Market", "Price",
            "Trend score", "Search query", "Link", "Image URL"]
    ws.append(cols)
    style_header(ws, len(cols))
    for it in items:
        ws.append([
            it["first_seen"], labels.get(it["category"], it["category"]),
            it["title"], it["source"], markets.get(it["country"], it["country"]),
            f'{it.get("currency") or ""} {it.get("price") or ""}'.strip(),
            it.get("trend_score"), it.get("query"), it["url"], it.get("image"),
        ])
        link = ws.cell(row=ws.max_row, column=9)
        if it["url"]:
            link.hyperlink, link.font = it["url"], LINK_FONT
    autosize(ws, [11, 22, 55, 12, 10, 12, 11, 26, 40, 40])

    # ---- Rising Searches ----
    ws = wb.create_sheet("Rising Searches")
    ws.append(["Market", "Category", "Rising query", "Growth"])
    style_header(ws, 4)
    for r in cat.get("rising_queries", []):
        ws.append([markets.get(r["market"], r["market"]),
                   labels.get(r["category"], r["category"]),
                   r["query"], r["growth"]])
    autosize(ws, [12, 22, 40, 12])

    # ---- Category Interest ----
    ws = wb.create_sheet("Category Interest")
    ws.append(["Market", "Category", "Search term", "Interest (0-100)"])
    style_header(ws, 4)
    for r in sorted(cat.get("category_interest", []),
                    key=lambda x: -x["score"]):
        ws.append([markets.get(r["market"], r["market"]),
                   labels.get(r["category"], r["category"]),
                   r["term"], r["score"]])
    autosize(ws, [12, 22, 34, 16])

    # ---- Source Status ----
    ws = wb.create_sheet("Source Status")
    ws.append(["Source", "Items fetched today", "Health"])
    style_header(ws, 3)
    for src, n in status.items():
        ws.append([src, n, "OK" if n > 0 or src == "google_trends" else "CHECK — returned 0"])
    autosize(ws, [18, 20, 24])

    path = OUT / f"GehnaRadar_{today}.xlsx"
    wb.save(path)
    shutil.copy(path, OUT / "latest.xlsx")
    print(f"Excel written: {path}")
    return path


if __name__ == "__main__":
    run()
